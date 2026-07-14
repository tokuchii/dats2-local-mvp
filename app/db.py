from __future__ import annotations

import csv
import io
import json
import logging
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator

import psycopg2
import psycopg2.pool
from openpyxl import Workbook, load_workbook

from .config import DATABASE_URL, MASTER_XLSX
from .sse import publish

log = logging.getLogger(__name__)

_pool: psycopg2.pool.ThreadedConnectionPool | None = None


def _get_pool() -> psycopg2.pool.ThreadedConnectionPool:
    global _pool
    if _pool is None:
        _pool = psycopg2.pool.ThreadedConnectionPool(1, 5, DATABASE_URL)
    return _pool


@contextmanager
def connect() -> Iterator[psycopg2.extensions.connection]:
    pool = _get_pool()
    conn = pool.getconn()
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        pool.putconn(conn)


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)


def _unjson(value: Any, default: Any) -> Any:
    if value is None:
        return default
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(value)
    except (json.JSONDecodeError, TypeError):
        return default


SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS systems (
    id                  SERIAL PRIMARY KEY,
    dats2_id            TEXT UNIQUE NOT NULL,
    name                TEXT NOT NULL,
    acronym             TEXT,
    developer_owner     TEXT,
    owner_type          TEXT,
    sector_commodity    TEXT,
    geographic_scope    TEXT,
    primary_category    TEXT NOT NULL,
    secondary_categories_json  JSONB NOT NULL DEFAULT '[]',
    commodity_tags_json        JSONB NOT NULL DEFAULT '[]',
    value_chain_tags_json      JSONB NOT NULL DEFAULT '[]',
    technology_tags_json       JSONB NOT NULL DEFAULT '[]',
    livestock_coverage  TEXT,
    core_function       TEXT,
    technology_channel  TEXT,
    primary_users       TEXT,
    maturity            TEXT,
    operating_status    TEXT,
    evidence_of_scale   TEXT,
    main_scaling_strength TEXT,
    primary_bottleneck  TEXT,
    interoperability    TEXT,
    interoperability_score INTEGER,
    source_url_1        TEXT,
    source_url_2        TEXT,
    evidence_confidence TEXT,
    sinag_priority      TEXT,
    recommended_sinag_action TEXT,
    payload_json        JSONB NOT NULL DEFAULT '{}',
    current_version     INTEGER NOT NULL DEFAULT 1,
    created_at          TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    updated_at          TIMESTAMPTZ NOT NULL DEFAULT NOW()
);

CREATE INDEX IF NOT EXISTS idx_systems_name ON systems (name);
CREATE INDEX IF NOT EXISTS idx_systems_primary_category ON systems (primary_category);
CREATE INDEX IF NOT EXISTS idx_systems_status ON systems (operating_status);
CREATE INDEX IF NOT EXISTS idx_systems_livestock ON systems (livestock_coverage);
CREATE INDEX IF NOT EXISTS idx_systems_dats2_id ON systems (dats2_id);

CREATE TABLE IF NOT EXISTS submissions (
    id              SERIAL PRIMARY KEY,
    source_type     TEXT NOT NULL,
    source_uri      TEXT,
    uploaded_path   TEXT,
    pasted_text     TEXT,
    submitted_by    TEXT,
    status          TEXT NOT NULL DEFAULT 'queued',
    error_message   TEXT,
    created_at      TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    updated_at      TIMESTAMPTZ NOT NULL DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS candidates (
    id              SERIAL PRIMARY KEY,
    submission_id   INTEGER NOT NULL REFERENCES submissions(id) ON DELETE CASCADE,
    status          TEXT NOT NULL DEFAULT 'proposed',
    assessment_mode TEXT NOT NULL,
    payload_json    JSONB NOT NULL DEFAULT '{}',
    evidence_json   JSONB NOT NULL DEFAULT '[]',
    duplicates_json JSONB NOT NULL DEFAULT '[]',
    confidence      DOUBLE PRECISION NOT NULL DEFAULT 0,
    reviewer_note   TEXT,
    created_at      TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    reviewed_at     TIMESTAMPTZ
);

CREATE TABLE IF NOT EXISTS system_versions (
    id              SERIAL PRIMARY KEY,
    system_id       INTEGER NOT NULL REFERENCES systems(id) ON DELETE CASCADE,
    version         INTEGER NOT NULL,
    payload_json    JSONB NOT NULL DEFAULT '{}',
    candidate_id    INTEGER REFERENCES candidates(id),
    created_at      TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    UNIQUE (system_id, version)
);

CREATE TABLE IF NOT EXISTS audit_events (
    id              SERIAL PRIMARY KEY,
    actor           TEXT NOT NULL,
    action          TEXT NOT NULL,
    entity_type     TEXT NOT NULL,
    entity_id       TEXT NOT NULL,
    details_json    JSONB NOT NULL DEFAULT '{}',
    created_at      TIMESTAMPTZ NOT NULL DEFAULT NOW()
);

CREATE INDEX IF NOT EXISTS idx_audit_created ON audit_events (created_at DESC);
"""


INVENTORY_KEYS = {
    "DATS2_ID": "dats2_id",
    "System / Tool": "name",
    "Acronym": "acronym",
    "Developer / Owner": "developer_owner",
    "Owner Type": "owner_type",
    "Sector / Commodity": "sector_commodity",
    "Geographic Scope": "geographic_scope",
    "DATS 2.0 Category": "primary_category",
    "Secondary DATS 2.0 Categories": "secondary_categories",
    "Commodity / Species Tags": "commodity_tags",
    "Value-chain Tags": "value_chain_tags",
    "Technology Tags": "technology_tags",
    "Livestock / Poultry Coverage": "livestock_coverage",
    "Core Function": "core_function",
    "Technology / Channel": "technology_channel",
    "Primary Users": "primary_users",
    "Maturity": "maturity",
    "Operating Status": "operating_status",
    "Evidence of Scale / Reach": "evidence_of_scale",
    "Main Scaling Strength": "main_scaling_strength",
    "Primary Bottleneck": "primary_bottleneck",
    "Interoperability / Data Governance": "interoperability",
    "Source URL 1": "source_url_1",
    "Source URL 2": "source_url_2",
    "Evidence Confidence": "evidence_confidence",
    "SINAG Priority": "sinag_priority",
    "Recommended SINAG Action": "recommended_sinag_action",
}


def split_tags(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    return [part.strip() for part in str(value).replace("|", ";").split(";") if part.strip()]


def init_db() -> None:
    with connect() as conn:
        with conn.cursor() as cur:
            cur.execute(SCHEMA_SQL)
            cur.execute("SELECT COUNT(*) FROM systems")
            count = cur.fetchone()[0]
    if count == 0 and MASTER_XLSX.exists():
        import_master_workbook(MASTER_XLSX)


def import_master_workbook(path: Path) -> int:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["DATS 2.0 Inventory"]
    headers = [str(c.value).strip() if c.value is not None else "" for c in sheet[4]]
    imported = 0
    timestamp = now_iso()
    with connect() as conn:
        cur = conn.cursor()
        for values in sheet.iter_rows(min_row=5, values_only=True):
            if not values or not values[0] or not values[1]:
                continue
            payload = {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i]}
            mapped: dict[str, Any] = {}
            for header, target in INVENTORY_KEYS.items():
                mapped[target] = payload.get(header)
            dats2_id = str(mapped["dats2_id"]).strip()
            cur.execute("SELECT 1 FROM systems WHERE dats2_id=%s", (dats2_id,))
            if cur.fetchone():
                continue
            secondary = split_tags(mapped.get("secondary_categories"))
            commodity = split_tags(mapped.get("commodity_tags"))
            value_chain = split_tags(mapped.get("value_chain_tags"))
            technology = split_tags(mapped.get("technology_tags"))
            cur.execute(
                """
                INSERT INTO systems (
                    dats2_id, name, acronym, developer_owner, owner_type, sector_commodity,
                    geographic_scope, primary_category, secondary_categories_json,
                    commodity_tags_json, value_chain_tags_json, technology_tags_json,
                    livestock_coverage, core_function, technology_channel, primary_users,
                    maturity, operating_status, evidence_of_scale, main_scaling_strength,
                    primary_bottleneck, interoperability, interoperability_score,
                    source_url_1, source_url_2, evidence_confidence, sinag_priority,
                    recommended_sinag_action, payload_json, current_version, created_at, updated_at
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING id
                """,
                (
                    dats2_id, str(mapped.get("name") or "").strip(), mapped.get("acronym"),
                    mapped.get("developer_owner"), mapped.get("owner_type"), mapped.get("sector_commodity"),
                    mapped.get("geographic_scope"), mapped.get("primary_category") or "Unclassified",
                    _json(secondary), _json(commodity), _json(value_chain), _json(technology),
                    mapped.get("livestock_coverage"), mapped.get("core_function"), mapped.get("technology_channel"),
                    mapped.get("primary_users"), mapped.get("maturity"), mapped.get("operating_status"),
                    mapped.get("evidence_of_scale"), mapped.get("main_scaling_strength"), mapped.get("primary_bottleneck"),
                    mapped.get("interoperability"), None, mapped.get("source_url_1"), mapped.get("source_url_2"),
                    mapped.get("evidence_confidence"), mapped.get("sinag_priority"), mapped.get("recommended_sinag_action"),
                    _json(payload), 1, timestamp, timestamp,
                ),
            )
            system_id = cur.fetchone()[0]
            cur.execute(
                "INSERT INTO system_versions(system_id, version, payload_json, created_at) VALUES (%s,%s,%s,%s)",
                (system_id, 1, _json(payload), timestamp),
            )
            imported += 1
        cur.execute(
            "INSERT INTO audit_events(actor,action,entity_type,entity_id,details_json,created_at) VALUES (%s,%s,%s,%s,%s,%s)",
            ("system", "import_master", "database", "systems", _json({"file": path.name, "imported": imported}), timestamp),
        )
    return imported


def row_to_system(row: tuple) -> dict[str, Any]:
    cols = [
        "id", "dats2_id", "name", "acronym", "developer_owner", "owner_type",
        "sector_commodity", "geographic_scope", "primary_category",
        "secondary_categories_json", "commodity_tags_json", "value_chain_tags_json",
        "technology_tags_json", "livestock_coverage", "core_function", "technology_channel",
        "primary_users", "maturity", "operating_status", "evidence_of_scale",
        "main_scaling_strength", "primary_bottleneck", "interoperability",
        "interoperability_score", "source_url_1", "source_url_2", "evidence_confidence",
        "sinag_priority", "recommended_sinag_action", "payload_json", "current_version",
        "created_at", "updated_at",
    ]
    item = dict(zip(cols, row))
    for key in ["secondary_categories_json", "commodity_tags_json", "value_chain_tags_json", "technology_tags_json"]:
        item[key.removesuffix("_json")] = _unjson(item.pop(key, None), [])
    item["payload"] = _unjson(item.pop("payload_json", None), {})
    if item.get("created_at"):
        item["created_at"] = item["created_at"].isoformat() if hasattr(item["created_at"], "isoformat") else str(item["created_at"])
    if item.get("updated_at"):
        item["updated_at"] = item["updated_at"].isoformat() if hasattr(item["updated_at"], "isoformat") else str(item["updated_at"])
    return item


def get_summary() -> dict[str, Any]:
    with connect() as conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM systems")
        systems = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM candidates WHERE status='proposed'")
        candidates = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM submissions WHERE status IN ('queued','running')")
        submissions = cur.fetchone()[0]
        cur.execute("SELECT primary_category, COUNT(*) FROM systems GROUP BY primary_category ORDER BY COUNT(*) DESC")
        categories = {row[0] or "Unclassified": row[1] for row in cur.fetchall()}
        cur.execute("SELECT operating_status, COUNT(*) FROM systems GROUP BY operating_status ORDER BY COUNT(*) DESC")
        statuses = {row[0] or "Unspecified": row[1] for row in cur.fetchall()}
        cur.execute("SELECT livestock_coverage, COUNT(*) FROM systems GROUP BY livestock_coverage ORDER BY COUNT(*) DESC")
        livestock = {row[0] or "Unspecified": row[1] for row in cur.fetchall()}
        cur.execute("SELECT COUNT(*) FROM systems WHERE jsonb_array_length(secondary_categories_json) > 0")
        multifunctional = cur.fetchone()[0]
    return {
        "systems": systems,
        "proposed_candidates": candidates,
        "queued_submissions": submissions,
        "categories": categories,
        "statuses": statuses,
        "livestock": livestock,
        "multifunctional": multifunctional,
    }


def get_filter_values() -> dict[str, list[str]]:
    with connect() as conn:
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT operating_status FROM systems WHERE operating_status IS NOT NULL ORDER BY operating_status")
        statuses = [r[0] for r in cur.fetchall()]
        cur.execute("SELECT DISTINCT livestock_coverage FROM systems WHERE livestock_coverage IS NOT NULL ORDER BY livestock_coverage")
        livestock = [r[0] for r in cur.fetchall()]
        cur.execute("""
            SELECT DISTINCT tag FROM systems,
            jsonb_array_elements_text(commodity_tags_json) AS tag
            WHERE tag IS NOT NULL ORDER BY tag
        """)
        commodities = [r[0] for r in cur.fetchall()]
    return {"statuses": statuses, "commodities": commodities, "livestock": livestock}


def list_systems(search: str = "", category: str = "", status: str = "", commodity: str = "", livestock: str = "", limit: int = 500, offset: int = 0) -> list[dict[str, Any]]:
    where: list[str] = []
    params: list[Any] = []
    if search:
        where.append("(name ILIKE %s OR acronym ILIKE %s OR developer_owner ILIKE %s OR core_function ILIKE %s)")
        term = f"%{search}%"
        params.extend([term, term, term, term])
    if category:
        where.append("(primary_category=%s OR secondary_categories_json::text ILIKE %s)")
        params.extend([category, f'%"{category}"%'])
    if status:
        where.append("operating_status=%s")
        params.append(status)
    if commodity:
        where.append("commodity_tags_json::text ILIKE %s")
        params.append(f"%{commodity}%")
    if livestock:
        where.append("livestock_coverage=%s")
        params.append(livestock)
    sql = "SELECT * FROM systems"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY (CASE WHEN dats2_id ~ '^D2-[0-9]+$' THEN CAST(SUBSTR(dats2_id, 4) AS INTEGER) ELSE 0 END) DESC LIMIT %s OFFSET %s"
    params.extend([min(max(limit, 1), 2000), max(offset, 0)])
    with connect() as conn:
        cur = conn.cursor()
        cur.execute(sql, params)
        return [row_to_system(row) for row in cur.fetchall()]


def count_systems(search: str = "", category: str = "", status: str = "", commodity: str = "", livestock: str = "") -> int:
    where: list[str] = []
    params: list[Any] = []
    if search:
        where.append("(name ILIKE %s OR acronym ILIKE %s OR developer_owner ILIKE %s OR core_function ILIKE %s)")
        term = f"%{search}%"
        params.extend([term, term, term, term])
    if category:
        where.append("(primary_category=%s OR secondary_categories_json::text ILIKE %s)")
        params.extend([category, f'%"{category}"%'])
    if status:
        where.append("operating_status=%s")
        params.append(status)
    if commodity:
        where.append("commodity_tags_json::text ILIKE %s")
        params.append(f"%{commodity}%")
    if livestock:
        where.append("livestock_coverage=%s")
        params.append(livestock)
    sql = "SELECT COUNT(*) FROM systems"
    if where:
        sql += " WHERE " + " AND ".join(where)
    with connect() as conn:
        cur = conn.cursor()
        cur.execute(sql, params)
        return cur.fetchone()[0]


def get_system(dats2_id: str) -> dict[str, Any] | None:
    with connect() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM systems WHERE dats2_id=%s", (dats2_id,))
        row = cur.fetchone()
        return row_to_system(row) if row else None


def create_submission(source_type: str, source_uri: str | None = None, uploaded_path: str | None = None, pasted_text: str | None = None, submitted_by: str | None = None) -> int:
    timestamp = now_iso()
    with connect() as conn:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO submissions(source_type,source_uri,uploaded_path,pasted_text,submitted_by,status,created_at,updated_at) VALUES (%s,%s,%s,%s,%s,'queued',%s,%s) RETURNING id",
            (source_type, source_uri, uploaded_path, pasted_text, submitted_by, timestamp, timestamp),
        )
        return cur.fetchone()[0]


def get_submission(submission_id: int) -> dict[str, Any] | None:
    with connect() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM submissions WHERE id=%s", (submission_id,))
        row = cur.fetchone()
        if not row:
            return None
        cols = [desc[0] for desc in cur.description]
        item = dict(zip(cols, row))
        for k in ("created_at", "updated_at"):
            if item.get(k):
                item[k] = item[k].isoformat() if hasattr(item[k], "isoformat") else str(item[k])
        return item


def update_submission(submission_id: int, *, status: str, error_message: str | None = None) -> None:
    with connect() as conn:
        cur = conn.cursor()
        cur.execute(
            "UPDATE submissions SET status=%s, error_message=%s, updated_at=%s WHERE id=%s",
            (status, error_message, now_iso(), submission_id),
        )
    publish("submission_status_changed", {"submission_id": submission_id, "status": status, "error_message": error_message})


def create_candidate(submission_id: int, payload: dict[str, Any], evidence: list[dict[str, Any]], duplicates: list[dict[str, Any]], confidence: float, assessment_mode: str) -> int:
    timestamp = now_iso()
    with connect() as conn:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO candidates(submission_id,status,assessment_mode,payload_json,evidence_json,duplicates_json,confidence,created_at)
            VALUES (%s,'proposed',%s,%s,%s,%s,%s,%s) RETURNING id
            """,
            (submission_id, assessment_mode, _json(payload), _json(evidence), _json(duplicates), float(confidence), timestamp),
        )
        candidate_id = cur.fetchone()[0]
        cur.execute("UPDATE submissions SET status='needs_review', updated_at=%s WHERE id=%s", (timestamp, submission_id))
    publish("submission_status_changed", {"submission_id": submission_id, "status": "needs_review", "candidate_id": candidate_id})
    return candidate_id


def count_candidates(status: str = "proposed") -> int:
    with connect() as conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM candidates WHERE status=%s", (status,))
        return cur.fetchone()[0]


def list_candidates(status: str = "proposed") -> list[dict[str, Any]]:
    with connect() as conn:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT c.id, c.submission_id, c.status, c.assessment_mode, c.payload_json,
                   c.evidence_json, c.duplicates_json, c.confidence, c.reviewer_note,
                   c.created_at, c.reviewed_at,
                   s.source_type, s.source_uri, s.submitted_by
            FROM candidates c JOIN submissions s ON s.id=c.submission_id
            WHERE c.status=%s ORDER BY c.created_at DESC
            """,
            (status,),
        )
        rows = cur.fetchall()
        cols = [desc[0] for desc in cur.description]
    result = []
    for row in rows:
        item = dict(zip(cols, row))
        item["payload"] = _unjson(item.pop("payload_json"), {})
        item["evidence"] = _unjson(item.pop("evidence_json"), [])
        item["duplicates"] = _unjson(item.pop("duplicates_json"), [])
        for k in ("created_at", "reviewed_at"):
            if item.get(k):
                item[k] = item[k].isoformat() if hasattr(item[k], "isoformat") else str(item[k])
        result.append(item)
    return result


def get_candidate(candidate_id: int) -> dict[str, Any] | None:
    with connect() as conn:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT c.id, c.submission_id, c.status, c.assessment_mode, c.payload_json,
                   c.evidence_json, c.duplicates_json, c.confidence, c.reviewer_note,
                   c.created_at, c.reviewed_at,
                   s.source_type, s.source_uri, s.uploaded_path, s.pasted_text, s.submitted_by
            FROM candidates c JOIN submissions s ON s.id=c.submission_id WHERE c.id=%s
            """,
            (candidate_id,),
        )
        row = cur.fetchone()
    if not row:
        return None
    cols = [
        "id", "submission_id", "status", "assessment_mode", "payload_json",
        "evidence_json", "duplicates_json", "confidence", "reviewer_note",
        "created_at", "reviewed_at",
        "source_type", "source_uri", "uploaded_path", "pasted_text", "submitted_by",
    ]
    item = dict(zip(cols, row))
    item["payload"] = _unjson(item.pop("payload_json"), {})
    item["evidence"] = _unjson(item.pop("evidence_json"), [])
    item["duplicates"] = _unjson(item.pop("duplicates_json"), [])
    for k in ("created_at", "reviewed_at"):
        if item.get(k):
            item[k] = item[k].isoformat() if hasattr(item[k], "isoformat") else str(item[k])
    return item


def save_candidate_payload(candidate_id: int, payload: dict[str, Any]) -> None:
    with connect() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE candidates SET payload_json=%s WHERE id=%s AND status='proposed'", (_json(payload), candidate_id))


def _payload_to_columns(payload: dict[str, Any]) -> dict[str, Any]:
    return {
        "name": payload.get("system_name") or "Unnamed system",
        "acronym": payload.get("acronym"),
        "developer_owner": payload.get("developer_owner"),
        "owner_type": payload.get("owner_type"),
        "sector_commodity": "; ".join(payload.get("sector_commodity", [])) if isinstance(payload.get("sector_commodity"), list) else payload.get("sector_commodity"),
        "geographic_scope": payload.get("geographic_scope"),
        "primary_category": payload.get("primary_category") or "Unclassified",
        "secondary_categories_json": _json(payload.get("secondary_categories", [])),
        "commodity_tags_json": _json(payload.get("commodity_species_tags", [])),
        "value_chain_tags_json": _json(payload.get("value_chain_tags", [])),
        "technology_tags_json": _json(payload.get("technology_tags", [])),
        "livestock_coverage": payload.get("livestock_poultry_coverage") or "Not dedicated",
        "core_function": payload.get("core_function"),
        "technology_channel": "; ".join(payload.get("technology_channel", [])) if isinstance(payload.get("technology_channel"), list) else payload.get("technology_channel"),
        "primary_users": "; ".join(payload.get("primary_users", [])) if isinstance(payload.get("primary_users"), list) else payload.get("primary_users"),
        "maturity": payload.get("maturity"),
        "operating_status": payload.get("operating_status"),
        "evidence_of_scale": payload.get("evidence_of_scale"),
        "main_scaling_strength": payload.get("main_scaling_strength"),
        "primary_bottleneck": payload.get("primary_bottleneck"),
        "interoperability": payload.get("interoperability_reason"),
        "interoperability_score": payload.get("interoperability_score"),
        "source_url_1": payload.get("source_url"),
        "source_url_2": payload.get("official_repository_url"),
        "evidence_confidence": payload.get("evidence_confidence"),
        "sinag_priority": payload.get("sinag_priority"),
        "recommended_sinag_action": payload.get("recommended_sinag_action"),
    }


def approve_candidate(candidate_id: int, actor: str, reviewer_note: str | None = None, merge_into: str | None = None) -> dict[str, Any]:
    candidate = get_candidate(candidate_id)
    if not candidate or candidate["status"] != "proposed":
        raise ValueError("Proposed candidate not found")
    payload = candidate["payload"]
    columns = _payload_to_columns(payload)
    timestamp = now_iso()
    with connect() as conn:
        cur = conn.cursor()
        if merge_into:
            cur.execute("SELECT * FROM systems WHERE dats2_id=%s", (merge_into,))
            existing = cur.fetchone()
            if not existing:
                raise ValueError("Merge target not found")
            cols = [desc[0] for desc in cur.description]
            existing_dict = dict(zip(cols, existing))
            system_id = existing_dict["id"]
            version = int(existing_dict["current_version"]) + 1
            set_clause = ", ".join(f"{key}=%s" for key in columns)
            cur.execute(
                f"UPDATE systems SET {set_clause}, payload_json=%s, current_version=%s, updated_at=%s WHERE id=%s",
                [*columns.values(), _json(payload), version, timestamp, system_id],
            )
            dats2_id = merge_into
        else:
            cur.execute(
                "SELECT MAX(CAST(SUBSTR(dats2_id, 4) AS INTEGER)) FROM systems WHERE dats2_id LIKE 'D2-%%' AND dats2_id NOT LIKE 'D2-WEB-%%'"
            )
            max_num = cur.fetchone()[0] or 0
            dats2_id = f"D2-{int(max_num)+1:03d}"
            cur.execute("SELECT 1 FROM systems WHERE dats2_id=%s FOR UPDATE", (dats2_id,))
            fields = ["dats2_id", *columns.keys(), "payload_json", "current_version", "created_at", "updated_at"]
            placeholders = ", ".join("%s" for _ in fields)
            cur.execute(
                f"INSERT INTO systems({','.join(fields)}) VALUES ({placeholders}) RETURNING id",
                [dats2_id, *columns.values(), _json(payload), 1, timestamp, timestamp],
            )
            system_id = cur.fetchone()[0]
            version = 1
        cur.execute(
            "INSERT INTO system_versions(system_id,version,payload_json,candidate_id,created_at) VALUES (%s,%s,%s,%s,%s)",
            (system_id, version, _json(payload), candidate_id, timestamp),
        )
        cur.execute(
            "UPDATE candidates SET status='approved', reviewer_note=%s, reviewed_at=%s WHERE id=%s",
            (reviewer_note, timestamp, candidate_id),
        )
        cur.execute(
            "UPDATE submissions SET status='approved', updated_at=%s WHERE id=%s",
            (timestamp, candidate["submission_id"]),
        )
        cur.execute(
            "INSERT INTO audit_events(actor,action,entity_type,entity_id,details_json,created_at) VALUES (%s,%s,%s,%s,%s,%s)",
            (actor, "approve_candidate", "system", dats2_id, _json({"candidate_id": candidate_id, "merge_into": merge_into, "note": reviewer_note}), timestamp),
        )
    return {"dats2_id": dats2_id, "version": version}


def reject_candidate(candidate_id: int, actor: str, reviewer_note: str | None = None) -> None:
    candidate = get_candidate(candidate_id)
    if not candidate or candidate["status"] != "proposed":
        raise ValueError("Proposed candidate not found")
    timestamp = now_iso()
    with connect() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE candidates SET status='rejected',reviewer_note=%s,reviewed_at=%s WHERE id=%s", (reviewer_note, timestamp, candidate_id))
        cur.execute("UPDATE submissions SET status='rejected',updated_at=%s WHERE id=%s", (timestamp, candidate["submission_id"]))
        cur.execute(
            "INSERT INTO audit_events(actor,action,entity_type,entity_id,details_json,created_at) VALUES (%s,%s,%s,%s,%s,%s)",
            (actor, "reject_candidate", "candidate", str(candidate_id), _json({"note": reviewer_note}), timestamp),
        )


def list_audit(limit: int = 200) -> list[dict[str, Any]]:
    with connect() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM audit_events ORDER BY id DESC LIMIT %s", (limit,))
        rows = cur.fetchall()
        cols = [desc[0] for desc in cur.description]
    result = []
    for row in rows:
        item = dict(zip(cols, row))
        item["details"] = _unjson(item.pop("details_json"), {})
        if item.get("created_at"):
            item["created_at"] = item["created_at"].isoformat() if hasattr(item["created_at"], "isoformat") else str(item["created_at"])
        result.append(item)
    return result


def export_xlsx_bytes() -> bytes:
    systems = list_systems(limit=5000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Systems"
    headers = [
        "DATS2_ID", "System / Tool", "Acronym", "Developer / Owner", "Owner Type",
        "Sector / Commodity", "Geographic Scope", "Primary Category", "Secondary Categories",
        "Commodity / Species Tags", "Value-chain Tags", "Technology Tags", "Livestock / Poultry Coverage",
        "Core Function", "Technology / Channel", "Primary Users", "Maturity", "Operating Status",
        "Evidence of Scale", "Primary Bottleneck", "Interoperability", "Interop Score",
        "Source URL 1", "Source URL 2", "Evidence Confidence", "SINAG Priority", "Recommended SINAG Action",
        "Version", "Updated At",
    ]
    ws.append(headers)
    for item in systems:
        ws.append([
            item["dats2_id"], item["name"], item.get("acronym"), item.get("developer_owner"), item.get("owner_type"),
            item.get("sector_commodity"), item.get("geographic_scope"), item.get("primary_category"),
            "; ".join(item.get("secondary_categories", [])), "; ".join(item.get("commodity_tags", [])),
            "; ".join(item.get("value_chain_tags", [])), "; ".join(item.get("technology_tags", [])),
            item.get("livestock_coverage"), item.get("core_function"), item.get("technology_channel"),
            item.get("primary_users"), item.get("maturity"), item.get("operating_status"), item.get("evidence_of_scale"),
            item.get("primary_bottleneck"), item.get("interoperability"), item.get("interoperability_score"),
            item.get("source_url_1"), item.get("source_url_2"), item.get("evidence_confidence"), item.get("sinag_priority"),
            item.get("recommended_sinag_action"), item.get("current_version"), item.get("updated_at"),
        ])
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    ws.freeze_panes = "A2"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_csv_bytes() -> bytes:
    systems = list_systems(limit=5000)
    fieldnames = [
        "dats2_id", "name", "acronym", "developer_owner", "owner_type",
        "sector_commodity", "geographic_scope", "primary_category",
        "secondary_categories", "commodity_tags", "value_chain_tags",
        "technology_tags", "livestock_coverage", "core_function",
        "technology_channel", "primary_users", "maturity", "operating_status",
        "evidence_of_scale", "primary_bottleneck", "interoperability",
        "interoperability_score", "source_url_1", "source_url_2",
        "evidence_confidence", "sinag_priority", "recommended_sinag_action",
        "current_version", "updated_at",
    ]
    out = io.StringIO()
    writer = csv.DictWriter(out, fieldnames=fieldnames)
    writer.writeheader()
    for item in systems:
        row = {}
        for key in fieldnames:
            val = item.get(key)
            if isinstance(val, list):
                val = "; ".join(str(v) for v in val)
            row[key] = val
        writer.writerow(row)
    return out.getvalue().encode("utf-8-sig")
