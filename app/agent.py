from __future__ import annotations

import ipaddress
import json
import logging
import re
import socket
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

import httpx
from bs4 import BeautifulSoup
from pydantic import BaseModel, Field, ValidationError
from pypdf import PdfReader
from rapidfuzz import fuzz

from .config import OPENAI_API_KEY, OPENAI_MODEL

log = logging.getLogger(__name__)
from .db import connect, create_candidate, get_submission, update_submission


DATS_CATEGORIES = [
    "Registry & digital identity",
    "Advisory & e-extension",
    "Farm management & decision support",
    "Climate, weather & early warning",
    "Precision agriculture, EO, drones & IoT",
    "Pest, disease & biosecurity",
    "Inputs, seed & nutrient tools",
    "Mechanization & equipment-as-a-service",
    "Markets, e-commerce & logistics",
    "Finance, insurance & payments",
    "Traceability, food safety & MRV",
    "Fisheries & aquaculture",
    "Data infrastructure, interoperability & dashboards",
    "Training, capacity & innovation hubs",
]

CATEGORY_KEYWORDS: dict[str, list[str]] = {
    "Registry & digital identity": ["registry", "registration", "identification", "identity", "beneficiary", "credential", "directory", "animal id"],
    "Advisory & e-extension": ["advisory", "extension", "knowledge", "recommendation", "chatbot", "training", "learning", "expert", "sms alert"],
    "Farm management & decision support": ["farm management", "record keeping", "farm record", "decision support", "scheduler", "production monitoring", "crop calendar", "planning"],
    "Climate, weather & early warning": ["climate", "weather", "forecast", "drought", "flood", "early warning", "rainfall", "agrometeor"],
    "Precision agriculture, EO, drones & IoT": ["precision", "sensor", "iot", "drone", "uav", "satellite", "remote sensing", "gis", "webgis", "automation", "computer vision"],
    "Pest, disease & biosecurity": ["pest", "disease", "biosecurity", "diagnostic", "surveillance", "outbreak", "animal health", "plant health"],
    "Inputs, seed & nutrient tools": ["seed", "fertilizer", "nutrient", "soil fertility", "germplasm", "breed", "feed"],
    "Mechanization & equipment-as-a-service": ["machinery", "mechanization", "equipment rental", "tractor", "harvester", "sprayer"],
    "Markets, e-commerce & logistics": ["market", "price", "buyer", "e-commerce", "marketplace", "logistics", "ordering", "cold chain", "trade"],
    "Finance, insurance & payments": ["credit", "loan", "finance", "insurance", "payment", "crowdfunding", "claims"],
    "Traceability, food safety & MRV": ["traceability", "blockchain", "ledger", "certification", "compliance", "food safety", "mrv", "chain of custody"],
    "Fisheries & aquaculture": ["fisher", "fish", "aquaculture", "vessel", "catch", "fishpond", "aquafarm"],
    "Data infrastructure, interoperability & dashboards": ["database", "dashboard", "analytics", "information system", "data platform", "api", "portal", "interoperability"],
    "Training, capacity & innovation hubs": ["training program", "academy", "learning site", "innovation hub", "capacity building", "incubator", "hackathon"],
}


class EvidenceItem(BaseModel):
    claim: str
    evidence_summary: str
    source_url: str | None = None
    evidence_type: str = "submitted"
    confidence: float = Field(ge=0, le=1)


class Assessment(BaseModel):
    system_name: str
    acronym: str | None = None
    developer_owner: str | None = None
    owner_type: str | None = None
    sector_commodity: list[str] = Field(default_factory=list)
    geographic_scope: str | None = None
    core_function: str
    technology_channel: list[str] = Field(default_factory=list)
    primary_users: list[str] = Field(default_factory=list)
    primary_category: str
    secondary_categories: list[str] = Field(default_factory=list)
    commodity_species_tags: list[str] = Field(default_factory=list)
    value_chain_tags: list[str] = Field(default_factory=list)
    technology_tags: list[str] = Field(default_factory=list)
    livestock_poultry_coverage: str = "Not dedicated"
    maturity: str | None = None
    operating_status: str | None = None
    evidence_of_scale: str | None = None
    main_scaling_strength: str | None = None
    primary_bottleneck: str | None = None
    public_source_availability: str | None = None
    official_repository_url: str | None = None
    public_license: str | None = None
    public_api_documentation: str | None = None
    machine_readable_export: str | None = None
    interoperability_score: int = Field(ge=0, le=4)
    interoperability_reason: str
    potential_sinag_role: str | None = None
    sinag_priority: str = "Medium"
    recommended_sinag_action: str | None = None
    evidence_confidence: str = "Medium"
    overall_confidence: float = Field(ge=0, le=1)
    source_url: str | None = None
    evidence: list[EvidenceItem] = Field(default_factory=list)


SYSTEM_PROMPT = f"""
You are a conservative DATS 2.0 evidence-assessment agent for Philippine digital agriculture.
Return only JSON matching the supplied schema. Extract only claims supported by the submitted source.
Do not invent scale, operating status, APIs, repositories, licenses, source-code availability, or users.

Functional categories: {DATS_CATEGORIES}

Rules:
- Assign one primary category and supported secondary categories.
- Separate operational systems, pilots, prototypes, legacy systems and proposals.
- Interoperability: 0 planned/concept; 1 interface only; 2 internal integration or portable export;
  3 documented external API/governed exchange; 4 official open source, public license, schemas,
  versioned API and portable export.
- Absence of a public repository means 'not publicly located', not proof of proprietary code.
- Flag dedicated swine, poultry, mixed livestock/poultry, livestock regulatory/animal health, or not dedicated.
- Evidence summaries must be short paraphrases with confidence scores.
""".strip()


def validate_public_url(url: str) -> None:
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.hostname:
        raise ValueError("Only public HTTP/HTTPS URLs are accepted")
    host = parsed.hostname.lower()
    if host in {"localhost", "localhost.localdomain"} or host.endswith(".local"):
        raise ValueError("Local/private URLs are blocked")
    try:
        infos = socket.getaddrinfo(host, parsed.port or (443 if parsed.scheme == "https" else 80), type=socket.SOCK_STREAM)
    except socket.gaierror as exc:
        raise ValueError(f"Cannot resolve host: {host}") from exc
    for info in infos:
        ip = ipaddress.ip_address(info[4][0])
        if ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_reserved or ip.is_multicast or ip.is_unspecified:
            raise ValueError("Local/private network targets are blocked")


def fetch_url(url: str) -> dict[str, Any]:
    validate_public_url(url)
    headers = {"User-Agent": "DATS2Local/1.0 (+human-reviewed digital agriculture registry)"}
    with httpx.Client(follow_redirects=True, timeout=35, headers=headers) as client:
        response = client.get(url)
        response.raise_for_status()
        if int(response.headers.get("content-length", "0") or 0) > 12_000_000:
            raise ValueError("Remote file exceeds the 12 MB acquisition limit")
        content_type = response.headers.get("content-type", "").lower()
        final_url = str(response.url)
        if "pdf" in content_type or final_url.lower().endswith(".pdf"):
            reader = PdfReader(BytesIO(response.content))
            text = "\n".join((page.extract_text() or "") for page in reader.pages)
            return {"url": final_url, "title": Path(urlparse(final_url).path).name, "text": text[:220_000], "links": [], "content_type": content_type}
        if "html" not in content_type:
            return {"url": final_url, "title": None, "text": response.text[:220_000], "links": [], "content_type": content_type}

    soup = BeautifulSoup(response.text, "html.parser")
    for tag in soup(["script", "style", "noscript", "svg", "nav"]):
        tag.decompose()
    title = soup.title.get_text(" ", strip=True) if soup.title else None
    text = "\n".join(line.strip() for line in soup.get_text("\n").splitlines() if line.strip())
    links: list[dict[str, str]] = []
    seen: set[str] = set()
    for a in soup.find_all("a", href=True):
        href = urljoin(final_url, a["href"])
        p = urlparse(href)
        if p.scheme not in {"http", "https"} or href in seen:
            continue
        seen.add(href)
        label = a.get_text(" ", strip=True)[:180]
        if any(token in (href + " " + label).lower() for token in ["github", "gitlab", "repository", "source code", "api", "developer", "documentation", "download", "apk", "license"]):
            links.append({"url": href, "label": label})
    return {"url": final_url, "title": title, "text": text[:220_000], "links": links[:60], "content_type": content_type}


def extract_file(path: str) -> dict[str, Any]:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".pdf":
        reader = PdfReader(str(file_path))
        text = "\n".join((page.extract_text() or "") for page in reader.pages)
        return {"url": None, "title": file_path.name, "text": text[:220_000], "links": [], "content_type": "application/pdf"}
    if suffix in {".xlsx", ".xls"}:
        from openpyxl import load_workbook
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"Sheet: {sheet_name}")
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join(str(v) if v is not None else "" for v in row)
                if row_text.strip(" |"):
                    lines.append(row_text)
        wb.close()
        text = "\n".join(lines)[:220_000]
        return {"url": None, "title": file_path.name, "text": text, "links": [], "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    if suffix in {".txt", ".md", ".csv", ".json"}:
        return {"url": None, "title": file_path.name, "text": file_path.read_text(errors="ignore")[:220_000], "links": [], "content_type": "text/plain"}
    raise ValueError("Unsupported file type")


def acquire_submission(submission: dict[str, Any]) -> dict[str, Any]:
    if submission["source_type"] == "url" and submission.get("source_uri"):
        return fetch_url(submission["source_uri"])
    if submission["source_type"] == "file" and submission.get("uploaded_path"):
        doc = extract_file(submission["uploaded_path"])
        doc["url"] = submission.get("source_uri")
        return doc
    if submission["source_type"] == "text" and submission.get("pasted_text"):
        return {"url": submission.get("source_uri"), "title": "Pasted system description", "text": submission["pasted_text"][:220_000], "links": [], "content_type": "text/plain"}
    raise ValueError("Submission has no usable source")


def sentences(text: str) -> list[str]:
    compact = re.sub(r"\s+", " ", text).strip()
    return [s.strip() for s in re.split(r"(?<=[.!?])\s+", compact) if len(s.strip()) >= 30]


def clean_title(title: str | None, text: str) -> str:
    candidate = (title or "").strip()
    for separator in [" | ", " – ", " — ", " :: "]:
        if separator in candidate:
            candidate = candidate.split(separator)[0].strip()
    generic = {"home", "homepage", "pasted system description", "untitled"}
    if not candidate or candidate.lower() in generic or len(candidate) < 4:
        first = next((line.strip() for line in text.splitlines() if 4 <= len(line.strip()) <= 140), "Submitted digital agriculture system")
        candidate = first
    return candidate[:180]


def infer_acronym(name: str, text: str) -> str | None:
    match = re.search(r"\(([A-Z][A-Z0-9-]{1,14})\)", name + " " + text[:1000])
    if match:
        return match.group(1)
    uppercase = re.findall(r"\b[A-Z][A-Z0-9-]{2,12}\b", name)
    blocked = {"THE", "AND", "FOR", "PHILIPPINES", "DOST", "DA"}
    return next((token for token in uppercase if token not in blocked), None)


def infer_developer(text: str, domain: str | None) -> str | None:
    patterns = [
        r"(?:developed|created|implemented|maintained|led) by\s+([^.;\n]{4,120})",
        r"(?:developer|implementing agency|project leader|owner)\s*[:\-]\s*([^.;\n]{4,120})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text[:20_000], flags=re.I)
        if match:
            return match.group(1).strip(" :-")[:180]
    return domain


def score_categories(text: str) -> dict[str, int]:
    lowered = text.lower()
    return {category: sum(lowered.count(keyword) for keyword in keywords) for category, keywords in CATEGORY_KEYWORDS.items()}


def infer_tags(text: str) -> tuple[list[str], list[str], list[str], str]:
    lower = text.lower()
    commodity = []
    for tag, keys in [
        ("Rice / palay", ["rice", "palay"]), ("Corn / maize", ["corn", "maize"]),
        ("Coconut", ["coconut"]), ("Banana", ["banana"]), ("Coffee", ["coffee"]),
        ("Cacao", ["cacao"]), ("Rubber", ["rubber"]), ("Vegetables", ["vegetable"]),
        ("Swine / native pig", ["swine", "pig", "sow", "piglet", "hog"]),
        ("Poultry / chicken / duck", ["poultry", "chicken", "duck", "itik", "egg incubator", "hatchery"]),
        ("Livestock — general", ["livestock", "animal health", "animal disease"]),
        ("Fisheries & aquaculture", ["fish", "fisher", "aquaculture", "vessel", "catch", "fishpond"]),
    ]:
        if any(key in lower for key in keys):
            commodity.append(tag)
    if not commodity:
        commodity = ["Agriculture — general"]

    tech = []
    for tag, keys in [
        ("Web / portal", ["web", "portal", "website"]), ("Mobile / Android", ["mobile", "android", "app"]),
        ("SMS / text", ["sms", "text message"]), ("Offline-capable", ["offline"]),
        ("Database / MIS", ["database", "information system"]), ("Dashboard / analytics", ["dashboard", "analytics"]),
        ("GIS / WebGIS", ["gis", "webgis", "geospatial", "mapping"]), ("Remote sensing / EO", ["remote sensing", "satellite"]),
        ("UAV / drone", ["uav", "drone"]), ("IoT / sensors", ["iot", "sensor"]),
        ("AI / ML / computer vision", ["artificial intelligence", "machine learning", "computer vision", "cnn", "image classifier", "large language model"]),
        ("Blockchain / DLT", ["blockchain", "distributed ledger"]), ("API / data exchange", [" api ", "application programming interface", "data exchange"]),
        ("QR / RFID / digital ID", ["qr code", "rfid"]),
    ]:
        if any(key in f" {lower} " for key in keys):
            tech.append(tag)

    value_chain = []
    for tag, keys in [
        ("Registration & identity", ["registry", "registration", "identification"]),
        ("R&D, breeding & genetic resources", ["research", "breed", "germplasm"]),
        ("Inputs, seed & feed", ["seed", "fertilizer", "feed", "nutrient"]),
        ("Production & farm management", ["production", "farm management", "record keeping", "monitoring"]),
        ("Animal/plant health & biosecurity", ["pest", "disease", "biosecurity", "diagnostic"]),
        ("Harvest & postharvest", ["harvest", "postharvest", "storage"]),
        ("Logistics & distribution", ["logistics", "delivery", "distribution", "transport"]),
        ("Markets & trade", ["market", "price", "buyer", "trade"]),
        ("Finance, insurance & payments", ["finance", "credit", "insurance", "payment"]),
        ("Traceability, compliance & MRV", ["traceability", "compliance", "certification", "mrv"]),
        ("Extension, training & support", ["extension", "training", "advisory", "expert"]),
    ]:
        if any(key in lower for key in keys):
            value_chain.append(tag)

    has_swine = "Swine / native pig" in commodity
    has_poultry = "Poultry / chicken / duck" in commodity
    has_livestock = "Livestock — general" in commodity
    if has_swine and has_poultry:
        coverage = "Dedicated — mixed livestock/poultry"
    elif has_swine:
        coverage = "Dedicated — swine/native pig"
    elif has_poultry:
        coverage = "Dedicated — poultry/chicken/duck"
    elif has_livestock and any(k in lower for k in ["animal health", "biosecurity", "surveillance", "regulatory"]):
        coverage = "Livestock regulatory / animal-health"
    elif has_livestock:
        coverage = "Livestock — general"
    else:
        coverage = "Not dedicated"
    return commodity, tech, value_chain, coverage


def heuristic_assessment(document: dict[str, Any]) -> Assessment:
    text = document.get("text", "")
    title = clean_title(document.get("title"), text)
    url = document.get("url")
    domain = urlparse(url).hostname if url else None
    scores = score_categories(title + "\n" + text[:80_000])
    ranked = sorted(scores, key=lambda c: scores[c], reverse=True)
    primary = ranked[0] if scores[ranked[0]] > 0 else "Data infrastructure, interoperability & dashboards"
    secondary = [c for c in ranked[1:] if scores[c] >= max(2, scores[primary] * 0.25)][:6]
    commodity, tech, value_chain, coverage = infer_tags(title + "\n" + text)
    lower = text.lower()
    status = "Status unverified"
    maturity = "Status requires review"
    if any(k in lower for k in ["launched", "operational", "currently available", "rollout", "deployed"]):
        status = "Active / status to verify"
        maturity = "Operational evidence reported"
    elif any(k in lower for k in ["prototype", "proof of concept", "capstone"]):
        status = "Research prototype"
        maturity = "Prototype"
    elif any(k in lower for k in ["proposed", "planned", "will develop", "under development"]):
        status = "Planned / development"
        maturity = "Concept / development"

    source_links = document.get("links", [])
    repository = next((item["url"] for item in source_links if any(k in item["url"].lower() for k in ["github.com", "gitlab.com"])), None)
    api_link = next((item["url"] for item in source_links if "api" in (item["url"] + " " + item.get("label", "")).lower()), None)
    export_evidence = None
    if any(k in lower for k in ["csv", "json export", "downloadable data", "export records", "excel export"]):
        export_evidence = "A portable export is mentioned; format and completeness require verification"
    public_source = "Official public repository located" if repository else "Public interface/source document found; official source repository not publicly located"
    if repository and api_link and export_evidence:
        interop_score = 3
    elif repository or api_link or export_evidence or any(k in lower for k in ["integrated system", "linked systems", "common database", "interoperability"]):
        interop_score = 2
    elif status.startswith("Planned"):
        interop_score = 0
    else:
        interop_score = 1

    all_sentences = sentences(text)
    scale_sentence = next((s for s in all_sentences if re.search(r"\b\d+[,.]?\d*\b", s) and any(k in s.lower() for k in ["farmer", "farm", "region", "province", "municip", "user", "site", "hectare", "national", "trained", "plantation"])), None)
    evidence: list[EvidenceItem] = []
    if all_sentences:
        evidence.append(EvidenceItem(claim="Core system description", evidence_summary=all_sentences[0][:450], source_url=url, confidence=0.65))
    if scale_sentence:
        evidence.append(EvidenceItem(claim="Scale or deployment evidence", evidence_summary=scale_sentence[:450], source_url=url, confidence=0.65))
    if repository:
        evidence.append(EvidenceItem(claim="Repository link", evidence_summary=f"The submitted page links to {repository}.", source_url=repository, evidence_type="repository", confidence=0.85))
    elif url:
        evidence.append(EvidenceItem(claim="Public source availability", evidence_summary="A public source page was retrieved, but no official repository link was detected.", source_url=url, confidence=0.75))

    users = []
    for label, keys in [
        ("Farmers", ["farmer"]), ("Extension workers", ["extension worker", "agricultural extension"]),
        ("Government agencies", ["department of agriculture", "government agency", "lgu"]),
        ("Researchers", ["researcher", "scientist"]), ("Agribusinesses", ["agribusiness", "company", "enterprise"]),
        ("Breeders", ["breeder"]), ("Fisherfolk", ["fisherfolk", "fishers"]),
    ]:
        if any(k in lower for k in keys):
            users.append(label)

    confidence = 0.55
    if url and len(text) > 1000:
        confidence += 0.1
    if scale_sentence:
        confidence += 0.05
    if repository or api_link:
        confidence += 0.05
    confidence = min(confidence, 0.78)
    evidence_confidence = "High" if confidence >= 0.75 else "Medium" if confidence >= 0.55 else "Low"

    return Assessment(
        system_name=title,
        acronym=infer_acronym(title, text),
        developer_owner=infer_developer(text, domain),
        owner_type=None,
        sector_commodity=commodity,
        geographic_scope="Philippines / scope requires verification" if any(k in lower for k in ["philippines", "philippine"]) else "Scope requires verification",
        core_function=(all_sentences[0][:700] if all_sentences else f"Digital agriculture system described as {title}."),
        technology_channel=tech,
        primary_users=users,
        primary_category=primary,
        secondary_categories=secondary,
        commodity_species_tags=commodity,
        value_chain_tags=value_chain,
        technology_tags=tech,
        livestock_poultry_coverage=coverage,
        maturity=maturity,
        operating_status=status,
        evidence_of_scale=scale_sentence,
        main_scaling_strength="A public description and identifiable institutional or technical evidence are available." if url else "A submitted description is available for review.",
        primary_bottleneck="Operational ownership, adoption evidence, data governance, interoperability and long-term maintenance require verification.",
        public_source_availability=public_source,
        official_repository_url=repository,
        public_license="Not publicly identified",
        public_api_documentation=api_link or "No public API documentation detected",
        machine_readable_export=export_evidence or "No machine-readable export evidence detected",
        interoperability_score=interop_score,
        interoperability_reason="Score is based on detected public interfaces, technical links, export statements and integration claims; reviewer verification is required.",
        potential_sinag_role="Catalog and referral; upgrade to a governed connector if identifiers, exports or APIs are verified.",
        sinag_priority="High" if primary in {"Registry & digital identity", "Climate, weather & early warning", "Pest, disease & biosecurity", "Data infrastructure, interoperability & dashboards"} else "Medium",
        recommended_sinag_action="Verify the owner, current operating status, source license, data model, export/API capability, privacy controls and maintenance plan before integration.",
        evidence_confidence=evidence_confidence,
        overall_confidence=confidence,
        source_url=url,
        evidence=evidence,
    )


def openai_assessment(document: dict[str, Any]) -> Assessment | None:
    if not OPENAI_API_KEY:
        return None
    try:
        from openai import OpenAI
        client = OpenAI(api_key=OPENAI_API_KEY)
        user_content = json.dumps({
            "source_url": document.get("url"),
            "title": document.get("title"),
            "technical_links": document.get("links", []),
            "content": document.get("text", "")[:110_000],
        }, ensure_ascii=False)
        response = client.chat.completions.create(
            model=OPENAI_MODEL,
            temperature=0,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT + "\n\nReturn JSON matching this schema:\n" + json.dumps(Assessment.model_json_schema(), indent=2)},
                {"role": "user", "content": user_content},
            ],
        )
        raw = response.choices[0].message.content
        result = Assessment.model_validate_json(raw)
        if not result.source_url:
            result.source_url = document.get("url")
        return result
    except Exception as exc:
        log.warning("OpenAI assessment failed, falling back: %s", exc)
        return None


def find_duplicates(name: str, acronym: str | None, limit: int = 5) -> list[dict[str, Any]]:
    with connect() as conn:
        cur = conn.cursor()
        cur.execute("SELECT dats2_id, name, acronym, developer_owner, primary_category FROM systems")
        rows = cur.fetchall()
    results = []
    for row in rows:
        dats2_id, sys_name, sys_acronym, developer_owner, primary_category = row
        name_score = fuzz.token_set_ratio(name, sys_name)
        acronym_score = fuzz.ratio((acronym or "").lower(), (sys_acronym or "").lower()) if acronym and sys_acronym else 0
        score = max(name_score, acronym_score)
        if score >= 52:
            results.append({
                "dats2_id": dats2_id, "name": sys_name, "acronym": sys_acronym,
                "developer_owner": developer_owner, "primary_category": primary_category,
                "score": round(score, 1),
            })
    return sorted(results, key=lambda item: item["score"], reverse=True)[:limit]


def process_submission(submission_id: int) -> int:
    update_submission(submission_id, status="running")
    try:
        submission = get_submission(submission_id)
        if not submission:
            raise ValueError("Submission not found")
        document = acquire_submission(submission)
        assessment = openai_assessment(document)
        if assessment:
            mode = f"OpenAI: {OPENAI_MODEL}"
        else:
            mode = "Deterministic fallback"
            assessment = heuristic_assessment(document)
        duplicates = find_duplicates(assessment.system_name, assessment.acronym)
        candidate_id = create_candidate(
            submission_id=submission_id,
            payload=assessment.model_dump(mode="json", exclude={"evidence"}),
            evidence=[item.model_dump(mode="json") for item in assessment.evidence],
            duplicates=duplicates,
            confidence=assessment.overall_confidence,
            assessment_mode=mode,
        )
        return candidate_id
    except Exception as exc:
        update_submission(submission_id, status="failed", error_message=str(exc)[:2000])
        raise
